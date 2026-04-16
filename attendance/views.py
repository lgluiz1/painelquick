import json
import string
import random
from django.http import JsonResponse
from django.views.decorators.csrf import csrf_exempt
from django.contrib.auth import authenticate, login, logout
from django.contrib.auth.models import User
from django.contrib.auth.decorators import login_required
from django.utils.text import slugify
from django.shortcuts import render, redirect, get_object_or_404
from django.db.models import Count, Q
from django.http import JsonResponse, HttpResponse
import os
import io
from django.template.loader import render_to_string
from openpyxl import Workbook
from xhtml2pdf import pisa
from .models import (
    Company, Branch, Meeting, Attendance, Course, Evaluation, Question, 
    StudentResponse, StudentAnswer, Complaint, 
    ComplaintCategory, UrgencyLevel, ComplaintUpdate, StaffProfile,
    Lead, GlobalConfig, GalleryImage, Testimonial, FAQItem
)
from functools import wraps
from django.core.exceptions import PermissionDenied
from livestream.youtube_service import YouTubeService
from livestream.models import YouTubeConfig, LiveEvent
from django.utils import timezone
from datetime import datetime
from django.conf import settings

# --- UTILITﾃヽIOS ---

def get_company_by_token(request):
    token = request.headers.get('X-Api-Key')
    if not token:
        return None
    return Company.objects.filter(api_token=token, is_active=True).first()

# Decorador de Controle de Acesso por Cargo
def staff_role_required(allowed_roles):
    def decorator(view_func):
        @wraps(view_func)
        @login_required
        def _wrapped_view(request, *args, **kwargs):
            # Superusuﾃ｡rios tﾃｪm acesso total bypassando o perfil
            if request.user.is_superuser:
                return view_func(request, *args, **kwargs)
            
            try:
                profile = request.user.staff_profile
                if profile.is_active and profile.role in allowed_roles:
                    return view_func(request, *args, **kwargs)
            except StaffProfile.DoesNotExist:
                pass
            
            from django.contrib import messages
            messages.error(request, "VOCE_NAO_TEM_PERMISSAO")
            return redirect('portal_dashboard')
        return _wrapped_view
    return decorator

# --- API Pﾃ咤LICA (COLETA DE DADOS - MULTI-TENANT) ---

@csrf_exempt
def submit_attendance(request):
    if request.method != 'POST':
        return JsonResponse({'error': 'POST required'}, status=405)
    
    company = get_company_by_token(request)
    if not company:
        return JsonResponse({'error': 'Token de Empresa Invﾃ｡lido ou Inativa'}, status=401)

    try:
        data = json.loads(request.body)
        name = data.get('nome')
        branch_name = data.get('filial')
        reuniao_title = data.get('reuniaoId', 'Geral')
        signature = data.get('assinatura')

        if not all([name, branch_name, reuniao_title, signature]):
            return JsonResponse({'error': 'Dados incompletos'}, status=400)

        meeting, _ = Meeting.objects.get_or_create(company=company, title=reuniao_title)
        Branch.objects.get_or_create(company=company, name=branch_name)

        Attendance.objects.create(
            meeting=meeting,
            name=name,
            branch=branch_name,
            signature=signature
        )
        return JsonResponse({'status': 'success', 'message': 'Presenﾃｧa registrada!'})
    except Exception as e:
        return JsonResponse({'error': str(e)}, status=500)

def api_get_active_evaluation(request, course_slug):
    company = get_company_by_token(request)
    if not company:
        return JsonResponse({'error': 'Token de Empresa Invﾃ｡lido'}, status=401)
        
    try:
        course = Course.objects.get(company=company, slug=course_slug)
        eval_active = course.evaluations.filter(is_active=True).first()
        
        if not eval_active:
            return JsonResponse({'error': 'Nenhuma avaliaﾃｧﾃ｣o ativa disponﾃｭvel'}, status=404)
        
        data = {
            'id': eval_active.id,
            'title': eval_active.title,
            'course': course.title,
            'questions': [{'id': q.id, 'text': q.text} for q in eval_active.questions.all()]
        }
        return JsonResponse(data)
    except Course.DoesNotExist:
        return JsonResponse({'error': 'Curso nﾃ｣o encontrado'}, status=404)

@csrf_exempt
def api_submit_evaluation(request):
    company = get_company_by_token(request)
    if not company:
        return JsonResponse({'error': 'Token de Empresa Invﾃ｡lido'}, status=401)

    try:
        data = json.loads(request.body)
        eval_id = data.get('evaluation_id')
        name = data.get('nome')
        branch = data.get('filial')
        email = data.get('email')
        answers = data.get('answers', {}) 
        
        evaluation = Evaluation.objects.get(pk=eval_id, course__company=company)
        response = StudentResponse.objects.create(evaluation=evaluation, name=name, branch=branch, email=email)
        
        for q_id, text in answers.items():
            question = Question.objects.get(pk=q_id, evaluation=evaluation)
            StudentAnswer.objects.create(response=response, question=question, answer_text=text)
            
        return JsonResponse({'status': 'success'})
    except Exception as e:
        return JsonResponse({'error': str(e)}, status=500)

@csrf_exempt
def api_get_complaint_status(request):
    company = get_company_by_token(request)
    if not company: return JsonResponse({'error': 'Token Invﾃ｡lido'}, status=401)
    
    ticket_id = request.GET.get('ticket_id')
    if not ticket_id: return JsonResponse({'error': 'Cﾃｳdigo obrigatﾃｳrio'}, status=400)
    
    try:
        complaint = company.complaints.get(ticket_id=ticket_id)
        last_update = complaint.updates.filter(is_from_admin=True).last()
        return JsonResponse({
            'status': complaint.get_status_display(),
            'created_at': complaint.created_at.strftime('%d/%m/%Y'),
            'response': last_update.message if last_update else "Seu relato estﾃ｡ sendo analisado com todo sigilo pelo Capelﾃ｣o."
        })
    except Complaint.DoesNotExist:
        return JsonResponse({'error': 'Cﾃｳdigo nﾃ｣o encontrado ou nﾃ｣o pertence a esta empresa.'}, status=404)

@csrf_exempt
def api_submit_complaint(request):
    company = get_company_by_token(request)
    if not company:
        return JsonResponse({'error': 'Token de Empresa Invﾃ｡lido'}, status=401)

    try:
        data = json.loads(request.body)
        is_anon = data.get('is_anonymous', True)
        branch = data.get('branch')
        description = data.get('description')

        if not all([branch, description]):
            return JsonResponse({'error': 'Dados incompletos'}, status=400)

        ticket = 'TK-' + ''.join(random.choices(string.ascii_uppercase + string.digits, k=6))
        category = ComplaintCategory.objects.filter(id=data.get('category_id'), company__isnull=True).first()
        urgency = UrgencyLevel.objects.filter(id=data.get('urgency_id'), company__isnull=True).first()

        Complaint.objects.create(
            company=company, ticket_id=ticket, is_anonymous=is_anon,
            name=data.get('name') if not is_anon else None,
            email=data.get('email') if not is_anon else None,
            branch=branch, category=category, urgency=urgency, description=description
        )
        return JsonResponse({'status': 'success', 'ticket_id': ticket})
    except Exception as e:
        return JsonResponse({'error': str(e)}, status=500)

@csrf_exempt
def api_get_branches(request):
    company = get_company_by_token(request)
    if not company:
        return JsonResponse({'error': 'Token Invﾃ｡lido'}, status=401)
    
    branches = company.branches.all().order_by('name')
    return JsonResponse({
        'branches': [{'id': b.id, 'name': b.name} for b in branches]
    })

@csrf_exempt
def api_complaint_options(request):
    company = get_company_by_token(request)
    if not company: return JsonResponse({'error': 'Invalid Token'}, status=401)
    return JsonResponse({
        'categories': [{'id': c.id, 'name': c.name} for c in ComplaintCategory.objects.filter(company__isnull=True)],
        'urgencies': [{'id': u.id, 'name': u.name, 'color': u.color} for u in UrgencyLevel.objects.filter(company__isnull=True)],
        'branches': [{'id': b.id, 'name': b.name} for b in company.branches.all()]
    })

# --- PORTAL DO CAPELﾃグ (VISTAS PRINCIPAIS) ---

@csrf_exempt
def portal_login(request):
    if request.method == 'POST':
        user = authenticate(username=request.POST.get('username'), password=request.POST.get('password'))
        if user:
            login(request, user)
            return redirect('portal_dashboard')
        return render(request, 'attendance/portal_login.html', {'error': 'Acesso negado.'})
    return render(request, 'attendance/portal_login.html')

def portal_logout(request):
    logout(request)
    return redirect('portal_login')

@staff_role_required(allowed_roles=['ADMIN'])
def portal_users(request):
    if request.method == 'POST':
        action = request.POST.get('action')
        
        if action == 'create_staff':
            username = request.POST.get('username')
            password = request.POST.get('password')
            role = request.POST.get('role')
            
            if User.objects.filter(username=username).exists():
                from django.contrib import messages
                messages.error(request, f"Usuﾃ｡rio {username} jﾃ｡ existe.")
            else:
                user = User.objects.create_user(username=username, password=password)
                StaffProfile.objects.create(user=user, role=role)
                from django.contrib import messages
                messages.success(request, f"Usuﾃ｡rio {username} criado com cargo {role}.")
        
        elif action == 'update_staff':
            profile_id = request.POST.get('profile_id')
            profile = get_object_or_404(StaffProfile, id=profile_id)
            profile.role = request.POST.get('role')
            profile.is_active = request.POST.get('is_active') == 'on'
            profile.save()
            from django.contrib import messages
            messages.success(request, f"Perfil de {profile.user.username} atualizado.")
            
        elif action == 'delete_staff':
            profile = get_object_or_404(StaffProfile, id=request.POST.get('profile_id'))
            username = profile.user.username
            profile.user.delete() # Deleta user e profile via cascade
            from django.contrib import messages
            messages.success(request, f"Usuﾃ｡rio {username} removido da equipe.")
            
        return redirect('portal_users')

    staff_members = StaffProfile.objects.all().select_related('user').order_by('-created_at')
    return render(request, 'attendance/portal_users.html', {
        'staff_members': staff_members,
        'roles': StaffProfile.ROLE_CHOICES
    })

@staff_role_required(allowed_roles=['ADMIN', 'AUDITOR'])
def portal_companies(request):
    if request.method == 'POST':
        # ... manter lﾃｳgica de criaﾃｧﾃ｣o/ediﾃｧﾃ｣o/delete existente ...
        action = request.POST.get('action')
        if action == 'create_company':
            name = request.POST.get('name')
            Company.objects.create(name=name, slug=request.POST.get('slug') or slugify(name))
        elif action == 'update_company':
            c = get_object_or_404(Company, id=request.POST.get('company_id'))
            c.name = request.POST.get('name')
            c.logo_url = request.POST.get('logo_url')
            c.primary_color = request.POST.get('primary_color')
            c.is_active = request.POST.get('is_active') == 'on'
            c.save()
        elif action == 'delete_company':
            get_object_or_404(Company, id=request.POST.get('company_id')).delete()
        return redirect('portal_companies')

    # Anotaﾃｧﾃｵes para a tabela
    companies = Company.objects.all().annotate(
        total_meetings=Count('meetings', distinct=True),
        total_complaints=Count('complaints', distinct=True),
        pending_complaints=Count('complaints', filter=Q(complaints__status='novo'), distinct=True)
    ).order_by('name')

    return render(request, 'attendance/portal_companies.html', {
        'companies': companies,
        'active_companies_count': Company.objects.filter(is_active=True).count(),
    })

def home(request):
    # --- Landing Page Pﾃｺblica ---
    config = GlobalConfig.get_solo()
    
    # 1. Logos de Empresas que confiam (Aleatﾃｳrios)
    trusted_companies = list(Company.objects.filter(is_active=True, logo_url__isnull=False).exclude(logo_url=""))
    random.shuffle(trusted_companies)
    trusted_companies = trusted_companies[:8]
    
    # 2. Galeria Dinﾃ｢mica
    gallery_images = GalleryImage.objects.filter(is_active=True)
    
    # 3. Depoimentos Moderados (Aleatﾃｳrios)
    testimonials = list(Testimonial.objects.filter(is_approved=True, show_on_home=True))
    random.shuffle(testimonials)
    
    # 4. Processamento de Lead
    if request.method == 'POST':
        first_name = request.POST.get('first_name')
        last_name = request.POST.get('last_name')
        email = request.POST.get('email')
        phone = request.POST.get('phone')
        company_name = request.POST.get('company')
        message = request.POST.get('message')
        
        Lead.objects.create(
            first_name=first_name, last_name=last_name, email=email,
            phone=phone, company=company_name, message=message
        )
        from django.contrib import messages
        messages.success(request, "Sua mensagem foi enviada! O Capelﾃ｣o entrarﾃ｡ em contato em breve.")
        return redirect('home')

    return render(request, 'attendance/home.html', {
        'config': config,
        'trusted_companies': trusted_companies,
        'gallery_images': gallery_images,
        'testimonials': testimonials,
    })

@staff_role_required(allowed_roles=['ADMIN', 'OUVIDOR', 'CONTEUDISTA', 'AUDITOR'])
def portal_dashboard(request):
    # --- Nova Dashboard Inovadora (Central de Alertas e Ranking) ---
    from livestream.models import LiveAttendance, LiveChatMessage
    from django.db.models import Sum, Max
    
    # 1. Ranking de Presenﾃｧa (Top 5 Empresas - ﾃ嗟timos 30 dias)
    thirty_days_ago = timezone.now() - timezone.timedelta(days=30)
    
    ranking = Company.objects.all().annotate(
        presencas_fisicas=Count('meetings__attendances', filter=Q(meetings__attendances__created_at__gte=thirty_days_ago), distinct=True),
        presencas_digitais=Count('live_events__attendances', filter=Q(live_events__attendances__timestamp__gte=thirty_days_ago, live_events__attendances__is_confirmed=True), distinct=True),
    )
    
    # Criar lista calculada para o Ranking
    company_ranking = []
    for c in ranking:
        score = c.presencas_fisicas + c.presencas_digitais
        company_ranking.append({'name': c.name, 'score': score})
    
    company_ranking = sorted(company_ranking, key=lambda x: x['score'], reverse=True)[:5]

    # 2. Alertas de Denﾃｺncias Nﾃ｣o Lidas
    unread_complaints = Complaint.objects.filter(is_read=False, status='novo').select_related('company').order_by('-created_at')[:10]
    
    # 3. Novos Leads do Site (Contatos)
    recent_leads = Lead.objects.filter(is_read=False).order_by('-created_at')[:5]

    # 4. Avaliaﾃｧﾃｵes com Prazo (Deadlines)
    approaching_evals = Evaluation.objects.filter(is_active=True, end_time__isnull=False).select_related('course__company').order_by('end_time')[:5]
    
    # 5. Mﾃｩtricas Globais Rﾃ｡pidas
    total_presences = Attendance.objects.count() + LiveAttendance.objects.count()

    context = {
        'company_ranking': company_ranking,
        'unread_complaints': unread_complaints,
        'recent_leads': recent_leads,
        'approaching_evals': approaching_evals,
        'total_presences_global': total_presences,
        'new_complaints_count': unread_complaints.count() + recent_leads.count(),
    }
    
    return render(request, 'attendance/portal_dashboard.html', context)

@staff_role_required(allowed_roles=['ADMIN', 'OUVIDOR', 'CONTEUDISTA', 'AUDITOR'])
def portal_company_detail(request, slug):
    company = get_object_or_404(Company, slug=slug)
    if request.method == 'POST':
        action = request.POST.get('action')
        if action == 'add_branch':
            Branch.objects.create(company=company, name=request.POST.get('branch_name'))
        elif action == 'update_settings':
            company.logo_url = request.POST.get('logo_url')
            company.primary_color = request.POST.get('primary_color')
            company.save()
        return redirect('portal_company_detail', slug=slug)
    return render(request, 'attendance/portal_company.html', {
        'company': company, 'meetings': company.meetings.all(),
        'courses': company.courses.all(), 'complaints': company.complaints.all(),
        'branches': company.branches.all()
    })

# --- GESTﾃグ DETALHADA ---

@staff_role_required(allowed_roles=['ADMIN', 'CONTEUDISTA'])
def portal_meetings(request, slug):
    company = get_object_or_404(Company, slug=slug)
    if request.method == 'POST':
        action = request.POST.get('action')
        if action == 'delete_meeting':
            get_object_or_404(Meeting, id=request.POST.get('meeting_id'), company=company).delete()
        elif action == 'create_meeting':
            title = request.POST.get('title')
            start_time_str = request.POST.get('start_time')
            description = request.POST.get('description', '')
            is_youtube_live = request.POST.get('is_youtube_live') == 'on'
            is_public = request.POST.get('is_public') == 'on'
            
            start_time = timezone.make_aware(datetime.strptime(start_time_str, '%Y-%m-%dT%H:%M'))
            
            # Criar a reuniﾃ｣o primeiro (como era antes)
            meeting = Meeting.objects.create(
                company=company,
                title=title,
                start_time=start_time,
                description=description,
                is_youtube_live=is_youtube_live   # <-- Faltava isso aqui!
            )
            
            if is_youtube_live:
                from livestream.youtube_service import YouTubeService
                from django.contrib import messages
                try:
                    yt = YouTubeService()
                    res = yt.create_live_broadcast(title, start_time, description)
                    if res:
                        from livestream.models import LiveEvent
                        live_event = LiveEvent.objects.create(
                            company=company,
                            title=title,
                            description=description,
                            youtube_url=res['youtube_url'],
                            start_time=start_time,
                            is_public=is_public
                        )
                        meeting.live_event = live_event
                        meeting.youtube_broadcast_id = res['broadcast_id']
                        meeting.youtube_stream_key = res['stream_key']
                        meeting.save()
                        messages.success(request, f"Live criada com sucesso no YouTube! ID: {res['broadcast_id']}")
                    else:
                        messages.error(request, "O YouTube nﾃ｣o retornou os dados da live. Verifique as configuraﾃｧﾃｵes.")
                except Exception as e:
                    messages.error(request, f"Erro crﾃｭtico ao vincular YouTube: {str(e)}")
        
        elif action == 'edit_meeting':
            meeting_id = request.POST.get('meeting_id')
            meeting = get_object_or_404(Meeting, id=meeting_id, company=company)
            
            title = request.POST.get('title')
            start_time_str = request.POST.get('start_time')
            description = request.POST.get('description', '')
            is_public = request.POST.get('is_public') == 'on'
            
            start_time = timezone.make_aware(datetime.strptime(start_time_str, '%Y-%m-%dT%H:%M'))
            
            # Update local Meeting
            meeting.title = title
            meeting.start_time = start_time
            meeting.description = description
            meeting.save()
            
            # Update associated LiveEvent
            if meeting.live_event:
                meeting.live_event.title = title
                meeting.live_event.description = description
                meeting.live_event.start_time = start_time
                meeting.live_event.is_public = is_public
                meeting.live_event.save()
                
                # Sync with YouTube if possible
                if meeting.youtube_broadcast_id:
                    from livestream.youtube_service import YouTubeService
                    yt = YouTubeService()
                    yt.update_live_broadcast(meeting.youtube_broadcast_id, title, start_time, description)

        return redirect('portal_meetings', slug=slug)
    return render(request, 'attendance/portal_meetings.html', {
        'company': company, 
        'meetings': company.meetings.all().order_by('-created_at'),
        'youtube_config': YouTubeConfig.get_solo()
    })

@staff_role_required(allowed_roles=['ADMIN', 'CONTEUDISTA'])
def portal_meeting_detail(request, slug, pk):
    meeting = get_object_or_404(Meeting, pk=pk, company__slug=slug)
    if request.method == 'POST':
        action = request.POST.get('action')
        if action == 'delete_attendance':
            get_object_or_404(Attendance, id=request.POST.get('attendance_id'), meeting=meeting).delete()
        elif action == 'update_live_status' and meeting.live_event:
            new_status = request.POST.get('status')
            if new_status in ['waiting', 'live', 'finished']:
                meeting.live_event.status = new_status
                meeting.live_event.save()
        elif action == 'confirm_attendances' and meeting.live_event:
            # Botﾃ｣o Master/Xerife: Confirma presenﾃｧa de quem estﾃ｡ ONLINE agora
            attendances = meeting.live_event.attendances.all()
            count = 0
            for la in attendances:
                if la.is_online():
                    la.is_confirmed = True
                    la.save()
                    count += 1
            # Poderﾃｭamos adicionar uma mensagem de sucesso aqui via messages framework
        return redirect('portal_meeting_detail', slug=slug, pk=pk)
    
    # Calcular estatﾃｭsticas da live se existir
    live_stats = {
        'total': 0,
        'online': 0,
        'offline': 0
    }
    if meeting.live_event:
        attendances = meeting.live_event.attendances.all()
        live_stats['total'] = attendances.count()
        for la in attendances:
            if la.is_online():
                live_stats['online'] += 1
            else:
                live_stats['offline'] += 1
                
    context = {
        'company': meeting.company, 
        'meeting': meeting, 
        'attendances': meeting.attendances.all().order_by('-created_at'),
        'live_stats': live_stats
    }
    return render(request, 'attendance/portal_meeting_detail.html', context)

@staff_role_required(allowed_roles=['ADMIN', 'CONTEUDISTA'])
def portal_courses(request, slug):
    company = get_object_or_404(Company, slug=slug)
    if request.method == 'POST':
        if request.POST.get('action') == 'delete_course':
            get_object_or_404(Course, id=request.POST.get('course_id'), company=company).delete()
        elif request.POST.get('action') == 'create_course':
            t = request.POST.get('title')
            Course.objects.create(company=company, title=t, slug=slugify(t))
        return redirect('portal_courses', slug=slug)
    return render(request, 'attendance/portal_courses.html', {'company': company, 'courses': company.courses.all().order_by('-created_at')})

@staff_role_required(allowed_roles=['ADMIN', 'CONTEUDISTA'])
def portal_course_detail(request, slug, pk):
    course = get_object_or_404(Course, pk=pk, company__slug=slug)
    if request.method == 'POST':
        if request.POST.get('action') == 'add_evaluation':
            Evaluation.objects.create(course=course, title=request.POST.get('title'))
        elif request.POST.get('action') == 'delete_evaluation':
            get_object_or_404(Evaluation, id=request.POST.get('evaluation_id'), course=course).delete()
        return redirect('portal_course_detail', slug=slug, pk=pk)
    return render(request, 'attendance/portal_course_detail.html', {'company': course.company, 'course': course, 'evaluations': course.evaluations.all().order_by('-created_at')})

@staff_role_required(allowed_roles=['ADMIN', 'OUVIDOR', 'AUDITOR'])
def portal_complaints(request, slug):
    company = get_object_or_404(Company, slug=slug)
    if request.method == 'POST' and request.POST.get('action') == 'delete_complaint':
        get_object_or_404(Complaint, id=request.POST.get('complaint_id'), company=company).delete()
        return redirect('portal_complaints', slug=slug)
    return render(request, 'attendance/portal_complaints.html', {'company': company, 'complaints': company.complaints.all().order_by('-created_at')})

@staff_role_required(allowed_roles=['ADMIN', 'OUVIDOR', 'AUDITOR'])
def portal_complaint_detail(request, slug, pk):
    complaint = get_object_or_404(Complaint, pk=pk, company__slug=slug)
    
    # Marcar como lida ao visualizar
    if not complaint.is_read:
        complaint.is_read = True
        complaint.save()

    if request.method == 'POST':
        action = request.POST.get('action')
        if action == 'add_update':
            ComplaintUpdate.objects.create(complaint=complaint, message=request.POST.get('message'), is_from_admin=True)
        elif action == 'update_status':
            complaint.status = request.POST.get('status')
            complaint.save()
        return redirect('portal_complaint_detail', slug=slug, pk=pk)
    return render(request, 'attendance/portal_complaint_detail.html', {'company': complaint.company, 'complaint': complaint, 'updates': complaint.updates.all().order_by('created_at')})

@login_required
def download_template(request, slug, feature):
    company = get_object_or_404(Company, slug=slug)
    
    # Caminho do arquivo na raiz do projeto (fora da pasta backend)
    # No Docker, a raiz pode estar em caminhos diferentes, mas tentaremos os padrﾃｵes
    file_map = {
        'presenca': 'presenca.html',
        'denuncia': 'denuncia.html',
        'avaliacao': 'avaliacao.html'
    }
    
    filename = file_map.get(feature)
    if not filename: return HttpResponse("Tipo invﾃ｡lido", status=400)
    
    # Tenta localizar o arquivo
    possible_paths = [
        f"/app/../{filename}", # Docker path (assuming backend is in /app)
        os.path.join(os.path.dirname(os.path.dirname(os.path.dirname(__file__))), filename) # Local path
    ]
    
    content = ""
    for path in possible_paths:
        if os.path.exists(path):
            with open(path, 'r', encoding='utf-8') as f:
                content = f.read()
            break
            
    if not content:
        return HttpResponse(f"Arquivo base {filename} nﾃ｣o encontrado no servidor.", status=404)
    
    # SUBSTITUIﾃ�髭S Mﾃ；ICAS:
    # 1. Token da Empresa
    content = content.replace('YOUR_X_API_KEY_HERE', company.api_token)
    content = content.replace('TOKEN_AQUI', company.api_token) # Variﾃ｡vel comum que podemos ter usado
    
    # 2. URL do Servidor (Baseado no seu domﾃｭnio atual)
    content = content.replace('http://localhost:8000', 'https://gilberto.luizgustavo.tech')
    content = content.replace('https://seu-servidor.com', 'https://gilberto.luizgustavo.tech')
    
    # 3. Branding (Opcional, se o HTML suportar)
    content = content.replace('#e63946', company.primary_color)
    if company.logo_url:
        content = content.replace('https://quickdelivery.com.br/wp-content/uploads/2025/09/logo-quick-delivery-solucoes-em-logistica.webp', company.logo_url)
    
    response = HttpResponse(content, content_type='text/html')
    response['Content-Disposition'] = f'attachment; filename="{feature}_{company.slug}.html"'
    return response

# --- Pﾃ；INAS HOSPEDADAS ---

def hosted_form(request, company_slug, feature):
    company = get_object_or_404(Company, slug=company_slug)
    context = {
        'company': company,
        'branches': company.branches.all()
    }
    
    if feature == 'presenca': 
        context['reuniao_id'] = request.GET.get('id', 'Geral')
    elif feature == 'denuncia' or feature == 'status_denuncia':
        context.update({
            'categories': ComplaintCategory.objects.filter(company__isnull=True),
            'urgencies': UrgencyLevel.objects.filter(company__isnull=True)
        })
    elif feature == 'avaliacao':
        # Carregado via JS no template para suportar o curso dinﾃ｢mico
        pass
        
    return render(request, f'attendance/public_{feature}.html', context)

@staff_role_required(allowed_roles=['ADMIN', 'CONTEUDISTA', 'AUDITOR'])
def portal_evaluation_detail(request, slug, pk):
    evaluation = get_object_or_404(Evaluation, pk=pk, course__company__slug=slug)
    
    if request.method == 'POST':
        action = request.POST.get('action')
        if action == 'add_question':
            text = request.POST.get('text')
            order = request.POST.get('order', 0)
            Question.objects.create(evaluation=evaluation, text=text, order=order)
        elif action == 'delete_question':
            get_object_or_404(Question, id=request.POST.get('question_id'), evaluation=evaluation).delete()
        elif action == 'toggle_status':
            evaluation.is_active = not evaluation.is_active
            evaluation.save()
        return redirect('portal_evaluation_detail', slug=slug, pk=pk)
        
    return render(request, 'attendance/portal_evaluation_detail.html', {
        'company': evaluation.course.company,
        'evaluation': evaluation,
        'questions': evaluation.questions.all(),
        'responses': evaluation.student_responses.all().order_by('-created_at')
    })

@staff_role_required(allowed_roles=['ADMIN', 'AUDITOR'])
def export_evaluation_excel(request, slug, pk):
    evaluation = get_object_or_404(Evaluation, pk=pk, course__company__slug=slug)
    responses = evaluation.student_responses.all().prefetch_related('answers__question')
    questions = evaluation.questions.all()

    wb = Workbook()
    ws = wb.active
    ws.title = "Resultados"

    # Header
    headers = ['Nome', 'Filial', 'Email', 'Data']
    for q in questions:
        headers.append(q.text)
    ws.append(headers)

    # Data
    for resp in responses:
        row = [resp.name, resp.branch, resp.email, resp.created_at.strftime('%d/%m/%Y %H:%M')]
        answers_map = {ans.question_id: ans.answer_text for ans in resp.answers.all()}
        for q in questions:
            row.append(answers_map.get(q.id, ""))
        ws.append(row)

    response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
    response['Content-Disposition'] = f'attachment; filename=resultados_{slugify(evaluation.title)}.xlsx'
    wb.save(response)
    return response

@login_required
def export_evaluation_pdf(request, slug, pk):
    evaluation = get_object_or_404(Evaluation, pk=pk, course__company__slug=slug)
    responses = evaluation.student_responses.all().prefetch_related('answers__question')
    questions = evaluation.questions.all()
    
    context = {
        'evaluation': evaluation,
        'responses': responses,
        'questions': questions,
        'company': evaluation.course.company,
    }
    
    html = render_to_string('attendance/pdf_evaluation_results.html', context)
    result = io.BytesIO()
    pdf = pisa.pisaDocument(io.BytesIO(html.encode("UTF-8")), result)
    
    if not pdf.err:
        response = HttpResponse(result.getvalue(), content_type='application/pdf')
        response['Content-Disposition'] = f'attachment; filename=resultados_{slugify(evaluation.title)}.pdf'
        return response
    return HttpResponse("Erro ao gerar PDF", status=500)

@staff_role_required(allowed_roles=['ADMIN', 'AUDITOR', 'OUVIDOR'])
def export_meeting_excel(request, slug, pk):
    meeting = get_object_or_404(Meeting, pk=pk, company__slug=slug)
    attendances = meeting.attendances.all().order_by('-created_at')

    wb = Workbook()
    ws = wb.active
    ws.title = "Lista de Presenﾃｧa"

    headers = ['Nome', 'Filial', 'Data de Registro']
    ws.append(headers)

    for att in attendances:
        ws.append([att.name, att.branch, att.created_at.strftime('%d/%m/%Y %H:%M')])

    response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
    response['Content-Disposition'] = f'attachment; filename=presenca_{slugify(meeting.title)}.xlsx'
    wb.save(response)
    return response

@staff_role_required(allowed_roles=['ADMIN', 'AUDITOR', 'OUVIDOR'])
def export_meeting_pdf(request, slug, pk):
    meeting = get_object_or_404(Meeting, pk=pk, company__slug=slug)
    attendances = meeting.attendances.all().order_by('name')
    
    context = {
        'meeting': meeting,
        'attendances': attendances,
        'company': meeting.company,
    }
    
    html = render_to_string('attendance/pdf_meeting_attendance.html', context)
    result = io.BytesIO()
    pdf = pisa.pisaDocument(io.BytesIO(html.encode("UTF-8")), result)
    
    if not pdf.err:
        response = HttpResponse(result.getvalue(), content_type='application/pdf')
        response['Content-Disposition'] = f'attachment; filename=presenca_{slugify(meeting.title)}.pdf'
        return response
    return HttpResponse("Erro ao gerar PDF", status=500)

@staff_role_required(allowed_roles=['ADMIN', 'CONTEUDISTA'])
def portal_saas_settings(request):
    from .models import FAQItem
    from livestream.models import YouTubeConfig
    from django.contrib import messages
    global_config = GlobalConfig.get_solo()
    
    # Verificação de arquivo físico na VPS
    import os
    from django.conf import settings
    secrets_file = os.path.join(settings.BASE_DIR, 'client_secret_337151481642-jne3vmg96u3jnghcm30t68tb4q18os97.apps.googleusercontent.com.json')
    has_secrets_file = os.path.exists(secrets_file)

    if request.method == 'POST':
        action = request.POST.get('action')
        
        if action == 'update_global':
            global_config.whatsapp_number = request.POST.get('whatsapp_number')
            global_config.contact_email = request.POST.get('contact_email')
            global_config.notify_email = request.POST.get('notify_email')
            global_config.address = request.POST.get('address')
            global_config.formation_year = request.POST.get('formation_year', 2004)
            global_config.save()
            messages.success(request, "Configurações Globais atualizadas!")
            
        elif action == 'add_faq':
            question = request.POST.get('question')
            answer = request.POST.get('answer')
            if question and answer:
                FAQItem.objects.create(question=question, answer=answer)
                messages.success(request, "Nova pergunta adicionada ao FAQ!")
            else:
                messages.error(request, "Preencha pergunta e resposta.")

        elif action == 'delete_faq':
            faq_id = request.POST.get('faq_id')
            FAQItem.objects.filter(id=faq_id).delete()
            messages.success(request, "Pergunta removida do FAQ.")

        elif action == 'add_category':
            ComplaintCategory.objects.create(name=request.POST.get('name'), company=None)
        elif action == 'delete_category':
            ComplaintCategory.objects.filter(id=request.POST.get('category_id'), company__isnull=True).delete()
        elif action == 'add_urgency':
            UrgencyLevel.objects.create(
                name=request.POST.get('name'), 
                color=request.POST.get('color', '#6c757d'),
                company=None
            )
        elif action == 'delete_urgency':
            UrgencyLevel.objects.filter(id=request.POST.get('urgency_id'), company__isnull=True).delete()
        
        return redirect('portal_saas_settings')

    context = {
        'categories': ComplaintCategory.objects.filter(company__isnull=True).order_by('name'),
        'urgencies': UrgencyLevel.objects.filter(company__isnull=True),
        'youtube_config': YouTubeConfig.get_solo(),
        'global_config': global_config,
        'has_secrets_file': has_secrets_file,
        'faqs': FAQItem.objects.all().order_by('order', '-id'),
    }
    return render(request, 'attendance/portal_saas_settings.html', context)

@staff_role_required(allowed_roles=['ADMIN'])
def youtube_auth(request):
    """Redireciona para o login do Google."""
    from django.urls import reverse
    import os
    
    # Forçamos HTTPS pois o domínio ja possui SSL ativo na VPS
    scheme = 'https'
    redirect_uri = f"{scheme}://{request.get_host()}{reverse('youtube_callback')}"
    
    try:
        flow = YouTubeService.get_flow(redirect_uri)
        authorization_url, state = flow.authorization_url(
            access_type='offline',
            include_granted_scopes='true',
            prompt='consent'
        )
        request.session['oauth_state'] = state
        # IMPORTANTE: Salvar o verifier para o PKCE não falhar no callback
        request.session['code_verifier'] = flow.code_verifier
        return redirect(authorization_url)
    except Exception as e:
        from django.contrib import messages
        messages.error(request, f"Erro ao iniciar autenticação: {str(e)}")
        return redirect('portal_saas_settings')

@staff_role_required(allowed_roles=['ADMIN'])
def youtube_callback(request):
    """Processa o retorno do Google e salva tokens."""
    from django.urls import reverse
    from django.contrib import messages
    import traceback
    
    state = request.session.get('oauth_state')
    scheme = 'https'
    redirect_uri = f"{scheme}://{request.get_host()}{reverse('youtube_callback')}"
    
    print(f"DEBUG YOUTUBE: Iniciando callback. State na sessão: {state}")
    
    try:
        flow = YouTubeService.get_flow(redirect_uri)
        
        # Recupera o verifier salvo no passo anterior
        flow.code_verifier = request.session.get('code_verifier')
        
        auth_response = request.build_absolute_uri()
        # Forçamos HTTPS na resposta para o flow.fetch_token não reclamar
        if 'http:' in auth_response and not settings.DEBUG:
            auth_response = auth_response.replace('http:', 'https:')
        elif 'gilberto.luizgustavo.tech' in auth_response:
            # Caso especial para o seu domínio
            auth_response = auth_response.replace('http:', 'https:')
            
        print(f"DEBUG YOUTUBE: Auth Response URI (Final): {auth_response}")
        
        flow.fetch_token(authorization_response=auth_response)
        
        creds = flow.credentials
        print(f"DEBUG YOUTUBE: Tokens obtidos com sucesso. Salvando no banco...")
        
        config = YouTubeConfig.get_solo()
        config.credentials = {
            'token': creds.token,
            'refresh_token': creds.refresh_token,
            'token_uri': creds.token_uri,
            'client_id': creds.client_id,
            'client_secret': creds.client_secret,
            'scopes': creds.scopes
        }
        config.is_active = True
        config.save()
        
        # Busca informações do canal para ficar bonitão no portal
        print(f"DEBUG YOUTUBE: Buscando informações do canal...")
        yt_service = YouTubeService()
        channels = yt_service.list_channels()
        if channels:
            channel = channels[0]
            config.channel_id = channel['id']
            config.channel_title = channel['snippet']['title']
            config.channel_thumbnail = channel['snippet']['thumbnails']['default']['url']
            config.save()
            messages.success(request, f"Sucesso! Canal '{config.channel_title}' vinculado.")
            print(f"DEBUG YOUTUBE: Canal '{config.channel_title}' vinculado com sucesso.")
        else:
            messages.warning(request, "Login realizado, mas nenhum canal do YouTube foi encontrado nesta conta.")
            print(f"DEBUG YOUTUBE: Nenhum canal encontrado.")
            
    except Exception as e:
        error_msg = f"Erro no callback do YouTube: {str(e)}"
        print(f"ERROR YOUTUBE: {error_msg}")
        traceback.print_exc()
        messages.error(request, error_msg)
        
    return redirect('portal_saas_settings')


@staff_role_required(allowed_roles=['ADMIN'])
def disconnect_youtube(request):
    """Desvincula a conta do YouTube."""
    from django.contrib import messages
    config = YouTubeConfig.get_solo()
    config.credentials = {}
    config.channel_id = None
    config.channel_title = None
    config.channel_thumbnail = None
    config.save()
    messages.info(request, "Serviﾃｧo do YouTube desvinculado com sucesso.")
    return redirect('portal_saas_settings')

def portal_logout(request):
    logout(request)
    return redirect('portal_login')

@staff_role_required(allowed_roles=['ADMIN', 'AUDITOR', 'OUVIDOR'])
def export_global_report_excel(request):
    companies = Company.objects.all().annotate(
        total_meetings=Count('meetings', distinct=True),
        total_complaints=Count('complaints', distinct=True),
        pending_complaints=Count('complaints', filter=Q(complaints__status='novo'), distinct=True)
    ).order_by('name')
    
    wb = Workbook()
    
    # Planilha 1: Dashboard Geral
    ws_dash = wb.active
    ws_dash.title = "Dashboard Geral"
    ws_dash.append(['RELATﾃ迭IO CONSOLIDADO - PORTAL DO CAPELﾃグ'])
    ws_dash.append(['Data de Geraﾃｧﾃ｣o:', timezone.now().strftime('%d/%m/%Y %H:%M')])
    ws_dash.append([])
    ws_dash.append(['Empresa', 'Reuniﾃｵes Totais', 'Total de Denﾃｺncias', 'Denﾃｺncias Pendentes'])
    
    for c in companies:
        ws_dash.append([c.name, c.total_meetings, c.total_complaints, c.pending_complaints])
        
    # Colunas auto-ajustﾃ｡veis bﾃ｡sicas
    for col in range(1, 5):
        ws_dash.column_dimensions[chr(64+col)].width = 20
    
    # Planilha 2: Todas as Denﾃｺncias Pendentes
    ws_comp = wb.create_sheet(title="Denﾃｺncias Pendentes")
    ws_comp.append(['Empresa', 'Ticket', 'Filial', 'Categoria', 'Urgﾃｪncia', 'Data'])
    
    pending_list = Complaint.objects.filter(status='novo').select_related('company', 'category', 'urgency').order_by('-created_at')
    for comp in pending_list:
        ws_comp.append([
            comp.company.name if comp.company else "Geral",
            comp.ticket_id,
            comp.branch,
            comp.category.name if comp.category else "Outra",
            comp.urgency.name if comp.urgency else "Normal",
            comp.created_at.strftime('%d/%m/%Y %H:%M')
        ])

    response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
    response['Content-Disposition'] = 'attachment; filename=relatorio_global_capelao.xlsx'
    wb.save(response)
    return response

# --- NOVAS FUNCIONALIDADES DINﾂMICAS ---

@staff_role_required(allowed_roles=['ADMIN', 'CONTEUDISTA'])
def portal_gallery(request):
    if request.method == 'POST':
        action = request.POST.get('action')
        if action == 'upload':
            files = request.FILES.getlist('images')
            for f in files:
                GalleryImage.objects.create(image=f)
        elif action == 'delete':
            img_id = request.POST.get('image_id')
            get_object_or_404(GalleryImage, id=img_id).delete()
        return redirect('portal_gallery')

    images = GalleryImage.objects.all()
    return render(request, 'attendance/portal_gallery.html', {'images': images})

@staff_role_required(allowed_roles=['ADMIN', 'CONTEUDISTA'])
def portal_testimonials(request):
    if request.method == 'POST':
        action = request.POST.get('action')
        test_id = request.POST.get('testimonial_id')
        testimonial = get_object_or_404(Testimonial, id=test_id)
        
        if action == 'approve':
            testimonial.is_approved = True
            testimonial.save()
        elif action == 'toggle_home':
            testimonial.show_on_home = not testimonial.show_on_home
            testimonial.save()
        elif action == 'delete':
            testimonial.delete()
        return redirect('portal_testimonials')

    testimonials = Testimonial.objects.all().order_by('-created_at')
    return render(request, 'attendance/portal_testimonials.html', {'testimonials': testimonials})

def public_feedback(request, company_slug):
    from django.contrib import messages
    company = get_object_or_404(Company, slug=company_slug)
    if request.method == 'POST':
        name = request.POST.get('name')
        role = request.POST.get('role')
        message = request.POST.get('message')
        photo = request.FILES.get('photo')
        
        try:
            Testimonial.objects.create(
                company_related=company,
                company_name=company.name,
                name=name,
                role=role,
                message=message,
                photo=photo
            )
            return render(request, 'attendance/public_feedback_success.html', {'company': company})
        except Exception as e:
            messages.error(request, f"Erro ao enviar depoimento: {str(e)}")
            # Retorna para a página com os dados para não perder o preenchimento
            return render(request, 'attendance/public_feedback.html', {
                'company': company,
                'name': name,
                'role': role,
                'message': message
            })
        
    return render(request, 'attendance/public_feedback.html', {'company': company})

def static_page(request, page_name):
    config = GlobalConfig.get_solo()
    template_map = {
        'quem-somos': 'attendance/page_about.html',
        'pacotes': 'attendance/page_packages.html',
        'duvidas': 'attendance/page_faq.html',
        'privacidade': 'attendance/page_privacy.html',
    }
    
    template = template_map.get(page_name)
    if not template:
        return redirect('home')
        
    context = {'config': config}
    
    if page_name == 'quem-somos':
        from django.utils import timezone
        from .models import Branch, Complaint, Attendance, Testimonial
        
        # 1. Anos de Experiência
        current_year = timezone.now().year
        context['experience_years'] = current_year - config.formation_year
        
        # 2. Empresas Atendidas (Filiais)
        context['companies_count'] = Branch.objects.count()
        
        # 3. Almas Impactadas (Soma de tudo)
        denuncias = Complaint.objects.count()
        presencas = Attendance.objects.count()
        depoimentos = Testimonial.objects.count()
        context['impacted_souls'] = denuncias + presencas + depoimentos

    if page_name == 'duvidas':
        context['faqs'] = FAQItem.objects.filter(is_active=True)
        
    return render(request, template, context)
